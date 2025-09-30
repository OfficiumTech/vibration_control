import os
import math
import matplotlib.pyplot as plt
import csv
from qgis.core import QgsProject, QgsVectorLayer, QgsFeature, QgsGeometry, QgsWkbTypes
from PyQt5.QtWidgets import QFileDialog, QTableWidgetItem, QDialog, QVBoxLayout, QLabel, QComboBox, QPushButton
import openpyxl
from openpyxl import Workbook
#import traceback QMessageBox,
from .processamento_sim import *


# Função principal chamada pelo botão Process
def processar_dados(iface, ui):
    try:
        # Obter os nomes das camadas selecionadas nos comboBoxes
        nome_furos = ui.comboBox_shpFuros.currentText()
        nome_geofones = ui.comboBox_shpGeofones.currentText()

        # Obter as camadas reais do projeto pelo nome
        camada_furos = obter_camada_por_nome(nome_furos)
        camada_geofones = obter_camada_por_nome(nome_geofones)

        if not camada_furos or not camada_geofones:
            raise Exception("Camadas de furos e geofones não encontradas. Verifique se foram selecionadas corretamente.")

        # Verificar se as camadas são 3D e armazenar mensagens de aviso caso não sejam
        avisos_3d = []
        z_field_furos = verificar_3d_ou_campo_z(camada_furos, nome_furos, avisos_3d)
        z_field_geofones = verificar_3d_ou_campo_z(camada_geofones, nome_geofones, avisos_3d)

        if avisos_3d:
            QMessageBox.warning(None, "Aviso - Dados 2D", "\n".join(avisos_3d))

        # Verificar e padronizar nomes de campos
        campos_furos = [campo.name() for campo in camada_furos.fields()]
        campos_geofones = [campo.name() for campo in camada_geofones.fields()]

        campo_seq = encontrar_nome_campo(campos_furos, ["Seq. D.", "Ret. (ms)", "seq. d.", "ret. (ms)", "Sequencia", "Sequência"])
        campo_carga = encontrar_nome_campo(campos_furos, ["Carga (KG)", "Cargas (KG)", "carga (kg)", "cargas (kg)", "Carga(KG)", "Cargas(KG)", "carga(kg)", "cargas(kg)" ])
        campo_id_furo = encontrar_nome_campo(campos_furos, ["Id", "id", "ID", "Identificação", "Ident", "Ponto", "PONTO"])
        campo_id_geofone = encontrar_nome_campo(campos_geofones, ["Id", "id", "ID", "Identificação", "Ident", "Ponto", "PONTO"])
        campo_ppv = encontrar_nome_campo(campos_geofones, ["PVS(mm/s)", "PPV(mm/s)", "PVS (mm/s)", "PPV (mm/s)", "PVS", "PPV"])
        desmonte_furo = encontrar_nome_campo(campos_furos, ["Desmonte", "desmonte", "Evento", "evento", "Plano de Fogo", "plano de fogo"])
        desmonte_geofone = encontrar_nome_campo(campos_geofones, ["Desmonte", "desmonte", "Evento", "evento", "Plano de Fogo", "plano de fogo"])

        if not all([campo_seq, campo_carga, campo_id_furo, campo_id_geofone, campo_ppv]):
            log("Campos obrigatórios ausentes na tabela de atributos. Leia a aba 'HELP'.", interromper=True)

        if desmonte_furo and not desmonte_geofone:
            log("Falta campo de identificação do desmonte na camada de geofones", interromper=True)
        if desmonte_geofone and not desmonte_furo:
            log("Falta campo de identificação do desmonte na camada de furos", interromper=True)

        evento_unico = False
        if not desmonte_geofone and not desmonte_furo:
            evento_unico = True

        # Mudar para a aba PROCESS
        ui.tabWidget.setCurrentIndex(1)

        # Agrupar furos por sequência de detonação em cada desmonte
        grupos_furos = {}
        for f in camada_furos.getFeatures():
            seq = f[campo_seq]
            carga = f[campo_carga]

            if not evento_unico:
                desmonte = f[desmonte_furo]
                chave = (desmonte, seq)
            else:
                chave = seq

            if chave not in grupos_furos:
                grupos_furos[chave] = []
            grupos_furos[chave].append((f, carga))

        # Calcular carga total por grupo
        cargas_por_seq = {chave: sum([c for _, c in furos]) for chave, furos in grupos_furos.items()}

        # Para cada desmonte, encontrar a(s) sequência(s) com maior carga
        sequencias_por_desmonte = {}
        for chave, carga in cargas_por_seq.items():
            if evento_unico:
                desmonte = None
                seq = chave
            else:
                desmonte, seq = chave

            if desmonte not in sequencias_por_desmonte:
                sequencias_por_desmonte[desmonte] = {"maior_carga": carga, "sequencias": [seq]}
            else:
                if carga > sequencias_por_desmonte[desmonte]["maior_carga"]:
                    sequencias_por_desmonte[desmonte] = {"maior_carga": carga, "sequencias": [seq]}
                elif carga == sequencias_por_desmonte[desmonte]["maior_carga"]:
                    sequencias_por_desmonte[desmonte]["sequencias"].append(seq)

        # Processar geofones
        resultados = []
        ui.resultados_processados = [] # Limpar resultados existentes
        for g in camada_geofones.getFeatures():
            geom_g = g.geometry()
            if geom_g.isEmpty() or QgsWkbTypes.geometryType(geom_g.wkbType()) != QgsWkbTypes.PointGeometry:
                continue

            pt_g = geom_g.constGet()
            xg, yg = pt_g.x(), pt_g.y()

            if QgsWkbTypes.hasZ(geom_g.wkbType()):
                zg = pt_g.z()
            else:
                zg = float(g[z_field_geofones]) if z_field_geofones else 0

            desmonte_atual = g[desmonte_geofone] if desmonte_geofone else None
            sequencias_validas = sequencias_por_desmonte.get(desmonte_atual, {}).get("sequencias", [])

            # Captura dos furos mais próximos
            menor_dist = float("inf")
            furo_mais_proximo = None
            seq_furo = None
            carga_furo = None
            xf_mais_proximo = None
            yf_mais_proximo = None
            zf_mais_proximo = None

            for chave, furos in grupos_furos.items():
                if evento_unico:
                    considerar = chave in sequencias_validas
                elif isinstance(chave, tuple):
                    desmonte_chave, seq_chave = chave
                    considerar = (desmonte_chave == desmonte_atual and seq_chave in sequencias_validas)
                else:
                    considerar = False

                if not considerar:
                    continue

                carga_total = cargas_por_seq[chave]

                for f, _ in furos:
                    geom_f = f.geometry()
                    if geom_f.isEmpty() or QgsWkbTypes.geometryType(geom_f.wkbType()) != QgsWkbTypes.PointGeometry:
                        continue

                    pt_f = geom_f.constGet()
                    xf, yf = pt_f.x(), pt_f.y()
                    if QgsWkbTypes.hasZ(geom_f.wkbType()):
                        zf = pt_f.z()
                    else:
                        zf = float(f[z_field_furos]) if z_field_furos else 0

                    dist = math.sqrt((xg - xf)**2 + (yg - yf)**2 + (zg - zf)**2)

                    if dist < menor_dist:
                        menor_dist = dist
                        furo_mais_proximo = f
                        seq_furo = chave[1] if isinstance(chave, tuple) else chave
                        carga_furo = carga_total
                        xf_mais_proximo = xf
                        yf_mais_proximo = yf
                        zf_mais_proximo = zf

            # Armazenar resultados
            if furo_mais_proximo:
                resultados.append({
                    "ID Geofone": g[campo_id_geofone],
                    "ID Furo": furo_mais_proximo[campo_id_furo],
                    "Seq. D.": seq_furo,
                    "Carga (KG)": carga_furo,
                    "Distância (m)": round(menor_dist, 2),
                    "PPV/PVS": g[campo_ppv],
                    "x do furo": xf_mais_proximo,
                    "y do furo": yf_mais_proximo,
                    "z do furo": zf_mais_proximo,
                    "Desmonte": desmonte_atual
                })

    

        processar_furoSim(ui)
            

        # Preencher a tabela da aba PROCESS com os resultados----------------------------------------
        ui.tabelaResultados.setRowCount(len(resultados))
        ui.tabelaResultados.setColumnCount(7)
        ui.tabelaResultados.setHorizontalHeaderLabels(["ID Geofone", "ID Furo", "Seq. D.", "Desmonte", "Carga (KG)", "Distância (m)", "PPV/PVS"])

        for i, res in enumerate(resultados):
            for j, chave in enumerate(["ID Geofone", "ID Furo", "Seq. D.", "Desmonte", "Carga (KG)", "Distância (m)", "PPV/PVS"]):
                item = QTableWidgetItem(str(res[chave]))
                ui.tabelaResultados.setItem(i, j, item)
        
        #ui.tabelaResultados.resizeColumnsToContents() # Ajustar automaticamente a largura das colunas
        # Definindo larguras fixas para cada coluna (em pixels)
        ui.tabelaResultados.setColumnWidth(0, 90)  # "ID Geofone"
        ui.tabelaResultados.setColumnWidth(1, 90)  # "ID Furo"
        ui.tabelaResultados.setColumnWidth(2, 70)   # "Seq. D."
        ui.tabelaResultados.setColumnWidth(3, 90)  # "Desmonte"
        ui.tabelaResultados.setColumnWidth(4, 100)   # "Carga (KG)"
        ui.tabelaResultados.setColumnWidth(5, 100)  # "Distância (m)"
        ui.tabelaResultados.setColumnWidth(6, 90)   # "PPV/PVS"

        #---------------------------------------------------------------------------------------------
        
        # Armazenar resultados para exportações posteriores
        ui.resultados_processados = resultados
        #ui.resultados_processadosFsim = resultadosFsim
        
        
    except Exception as e:
        tb = traceback.format_exc()
        log(f"⚠️ Erro no Processamento: {e} \nTraceback:\n{tb}")        
        #QMessageBox.critical(None, "Processamento do Vibration Control - ERRO!", str(e) "\nTraceback:\n"{tb})



# Função que retorna a camada do projeto com o nome informado
def obter_camada_por_nome(nome):
    for camada in QgsProject.instance().mapLayers().values():
        if camada.name() == nome:
            return camada
    return None


# Encontra o nome real de um campo da camada a partir de possíveis nomes
def encontrar_nome_campo(campos, possiveis_nomes):
    for nome in possiveis_nomes:
        if nome in campos:
            return nome
    return None


# Verifica se a camada é 3D ou tenta encontrar campo Z na tabela de atributos
def verificar_3d_ou_campo_z(camada, nome_camada, avisos):
    if QgsWkbTypes.hasZ(camada.wkbType()):
        return None  # Geometria já é 3D, não precisa de campo Z

    # Lista de nomes padrão de campos Z
    nomes_possiveis = ["z", "Z", "Cota", "Elevação"]
    campos = [campo.name() for campo in camada.fields()]

    for nome in nomes_possiveis:
        if nome in campos:
            avisos.append(f"A camada '{nome_camada}' é 2D. Usando campo '{nome}' como cota Z.")
            return nome

    # Nenhum campo padrão encontrado, solicitar ao usuário
    campo_escolhido = dialogo_escolher_campo_z(camada, nome_camada)
    if campo_escolhido:
        avisos.append(f"A camada '{nome_camada}' é 2D. Usando campo '{campo_escolhido}' como cota Z.")
    else:
        avisos.append(f"A camada '{nome_camada}' é 2D. Nenhum campo Z será usado.")
    return campo_escolhido


# Cria um popup para o usuário escolher um campo Z ou seguir sem cota
def dialogo_escolher_campo_z(camada, nome_camada):
    dialog = QDialog()
    dialog.setWindowTitle(f"Selecionar campo Z - {nome_camada}")
    layout = QVBoxLayout()

    label = QLabel("A geometria desta camada é 2D. Selecione o campo da tabela que representa a cota Z:")
    layout.addWidget(label)

    combo = QComboBox()
    campos = [campo.name() for campo in camada.fields()]
    combo.addItems(campos)
    combo.addItem("Sem cota z")
    layout.addWidget(combo)

    botao_ok = QPushButton("OK")
    botao_ok.clicked.connect(dialog.accept)
    layout.addWidget(botao_ok)

    dialog.setLayout(layout)

    if dialog.exec_() == QDialog.Accepted:
        escolhido = combo.currentText()
        if escolhido == "Sem cota z":
            return None
        return escolhido
    return None


# Exporta os resultados para txt, csv ou excel
def exportar_tabela_para_txt(resultados):
    caminho, filtro = QFileDialog.getSaveFileName(
        None,
        "Salvar resultados",
        "",
        "Texto (*.txt);;CSV (*.csv);;Excel (*.xlsx)"
    )

    if not caminho:
        return

    extensao = os.path.splitext(caminho)[1].lower()
    if extensao not in ['.txt', '.csv', '.xlsx']:
        QMessageBox.warning(None, "Aviso", "A extensão deve ser .txt, .csv ou .xlsx")
        return

    try:
        if extensao == ".xlsx":
            #wb = openpyxl.Workbook()
            wb = Workbook()
            ws = wb.active
            ws.append(["ID Geofone", "ID Furo", "Seq. D.", "Carga (KG)", "Distância (m)", "PPV/PVS"])
            for res in resultados:
                ws.append([
                    res['ID Geofone'], res['ID Furo'], res['Seq. D.'],
                    res['Carga (KG)'], res['Distância (m)'], res['PPV/PVS']
                ])
            wb.save(caminho)

        else:
            with open(caminho, "w", newline='', encoding="utf-8") as f:
                escritor = csv.writer(f, delimiter='\t')
                escritor.writerow(["ID Geofone", "ID Furo", "Seq. D.", "Carga (KG)", "Distância (m)", "PPV/PVS"])
                for res in resultados:
                    escritor.writerow([
                        res['ID Geofone'], res['ID Furo'], res['Seq. D.'],
                        res['Carga (KG)'], res['Distância (m)'], res['PPV/PVS']
                    ])

        QMessageBox.information(None, "Sucesso", f"Arquivo salvo com sucesso em:\n{caminho}")

    except Exception as e:
        tb = traceback.format_exc()
        log(f"ERRO ao exportar dados: {e}\nTraceback:\n{tb}")
        #QMessageBox.critical(None, "Erro", str(e))


#Logs
import traceback
from PyQt5.QtWidgets import QMessageBox

def log(mensagem=None, excecao=None, interromper=False):
    if excecao:
        mensagem_erro = f"{mensagem or 'Erro durante o processamento:'}\n\n{str(excecao)}\n\n{traceback.format_exc()}"
    else:
        mensagem_erro = mensagem or "Erro desconhecido."

    QMessageBox.information(None, "Processamento - Blaster Vibration Control", mensagem_erro)

    if interromper:
        raise Exception("Processo interrompido!")

        
#def log(mensagem):
    #QMessageBox.information(None, "Procesamento - Blaster Vibration Control", mensagem)
    
    
# Gera gráficos baseados nos resultados (PPV vs Distância e Carga vs Distância)
def gerar_grafico(resultados):
    try:
        distancias = [r['Distância (m)'] for r in resultados]
        ppvs = [r['PPV/PVS'] for r in resultados]

        plt.figure(figsize=(8, 5))
        plt.scatter(distancias, ppvs, color='darkgreen')
        plt.xlabel("Distância (m)")
        plt.ylabel("PPV/PVS")
        plt.title("PPV/PVS vs Distância")
        plt.grid(True)
        plt.tight_layout()
        plt.show()

        cargas = [r['Carga (KG)'] for r in resultados]
        plt.figure(figsize=(8, 5))
        plt.scatter(distancias, cargas, color='darkred')
        plt.xlabel("Distância (m)")
        plt.ylabel("Carga (KG)")
        plt.title("Carga vs Distância")
        plt.grid(True)
        plt.tight_layout()
        plt.show()

    except Exception as e:
        QMessageBox.critical(None, "Erro ao gerar gráficos", str(e))



def importar_resultados_excel(ui):
    """
    Importa dados de uma planilha Excel (.xlsx), preenche a variável ui.resultados_processados
    e atualiza a tabelaResultados com os dados importados.
    """
    try:
        # Abrir diálogo para selecionar arquivo
        caminho_arquivo, _ = QFileDialog.getOpenFileName(None, "Selecionar arquivo Excel", "", "Arquivos Excel (*.xlsx)")
        if not caminho_arquivo:
            return  # Usuário cancelou

        # Abrir planilha
        wb = openpyxl.load_workbook(caminho_arquivo, data_only=True)
        planilha = wb.active

        # Limpar resultados existentes
        ui.resultados_processados = []

        # Limpar tabelaResultados
        tabela = ui.tabelaResultados
        tabela.setRowCount(0)
        tabela.setColumnCount(0)

        # Ler cabeçalhos esperados
        cabecalhos_esperados = ["ID Geofone", "Carga (KG)", "Distância (m)", "PPV/PVS"]
        cabecalhos = [cell.value for cell in next(planilha.iter_rows(min_row=1, max_row=1))]

        # Verifica se os cabeçalhos estão corretos
        if not all(cab in cabecalhos for cab in cabecalhos_esperados):
            raise Exception("Cabeçalhos da planilha inválidos. Esperado: " + ", ".join(cabecalhos_esperados))

        # Mapear os índices dos cabeçalhos
        indices = {cab: cabecalhos.index(cab) for cab in cabecalhos_esperados}

        # Ler dados
        for linha in planilha.iter_rows(min_row=2):
            if all(linha[indices[cab]] is None for cab in cabecalhos_esperados):
                continue  # Ignora linhas vazias

            resultado = {
                "ID Geofone": str(linha[indices["ID Geofone"]].value).strip(),
                "Carga (KG)": float(linha[indices["Carga (KG)"]].value),
                "Distância (m)": float(linha[indices["Distância (m)"]].value),
                "PPV/PVS": float(linha[indices["PPV/PVS"]].value)
            }

            ui.resultados_processados.append(resultado)

        # Atualizar a tabelaResultados
        tabela.setColumnCount(len(cabecalhos_esperados))
        tabela.setHorizontalHeaderLabels(cabecalhos_esperados)
        tabela.setRowCount(len(ui.resultados_processados))

        for row, resultado in enumerate(ui.resultados_processados):
            for col, chave in enumerate(cabecalhos_esperados):
                item = QTableWidgetItem(str(resultado[chave]))
                #item.setFlags(item.flags() ^ Qt.ItemIsEditable)
                tabela.setItem(row, col, item)

        QMessageBox.information(None, "Processamento.py: Importação concluída", "Planilha importada com sucesso!")

    except Exception as e:
        tb = traceback.format_exc()
        QMessageBox.critical(None, "Processamento.py: Erro ao importar", f"Ocorreu um erro ao importar a planilha:\n{str(e)}\nTraceback:\n{tb}")